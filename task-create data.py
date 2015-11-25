from util import *
from openpyxl import *
from parser import *

p = Parser()

# ======== Get intersection of all titles ========

all_names = []
for filename in listdir(p.path) :
    same_name = []
    sheet     = p.load_workbook(filename).get_sheet_by_name('Combined Data')
    data = p.fetch_data(sheet.rows)
    all_names.append(data[0])

final_names = all_names[0]
for names in all_names[1:] :
    for name in list(final_names) :
        if name not in names :
            final_names.remove(name)
# print j(final_names)
# print len(final_names)

# ======== Create new xlsx containing all data ========

wb = Workbook()
sheet1 = wb.active
sheet1.title = 'All Data'
for i in range(len(final_names)):
    sheet1[chr(i + 65) + '1'] = final_names[i]
r = 2 # insert from row 2
for filename in listdir(p.path) :
    sheet    = p.load_workbook(filename).get_sheet_by_name('Combined Data')
    data     = p.sheet_to_data(sheet)
    # print len(data), len(data[0])
    for row in range(len(data)) :
        for index, name in enumerate(final_names) :
            sheet1[chr(index + 65) + str(r)] = data[row][name]
        r += 1

wb.save('../data/A/all_data.xlsx')