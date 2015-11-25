from util import *
from openpyxl import *

class Parser :
    path = '../data/A'

    def __init__(self):
        pass

    def load_workbook(self, filename):
        wb = load_workbook(join(self.path, filename))
        return wb

    def fetch_data(self, rows) :
        return [[cell.value for cell in row] for row in rows]

    def load_data(self, data, index = None) :
        fields = data[0]
        if index == None :
            _ = []
            for row in data[1:] :
                _.append(dict(zip(fields, row)))
            return _
        else :
            _ = {}
            for row in data[1:] :
                __ = dict(zip(fields, row))
                _[__[index]] = __
            return _

    def sheet_to_data(self, sheet) :
        data = []
        for row in sheet.rows[1:] :
            row_data = {}
            for index, cell in enumerate(row) :
                title_index = chr(index + 65)+'1'
                row_data[sheet[title_index].value] = cell.value
            data.append(row_data)
        # print j(data)
        return data

    def has_next_row(self, sheet, row_num) :
        # better to check the whole row 
        if str(sheet['A' + str(row_num + 1)].value) == 'None' :
            return False
        return True

    def has_next_col(self, sheet, col_num) :
        # better to check the whole col 
        if str(sheet[chr(col_num + 65) + '1'].value) == 'None' :
            return False
        return True

    def del_row(self, sheet, row_num) :
        while p.has_next_row(sheet, row_num - 1) :
            col_num = 0
            while p.has_next_col(sheet, col_num) :
                sheet[chr(col_num + 65) + str(row_num)] = sheet[chr(col_num + 65) + str(row_num + 1)].value
                col_num += 1
                # print chr(col_num + 65) + str(row_num) + ' to ' + chr(col_num + 65) + str(row_num + 1), col_num
            row_num += 1
        return sheet

    def create_test_wb(self, row, col, sheetname, filename) :
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = sheetname
        for i in range(row):
            for j in range(col):
                sheet1[chr(i + 65) + str(j+1)] = chr(i + 65) + str(j+1)
        wb.save(filename)
        return 